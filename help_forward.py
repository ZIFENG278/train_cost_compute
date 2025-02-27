
# how to generate forward graph and make it into table format
import json 
import re
from collections import defaultdict
import math
import numpy as np
from ops import *
from module import *
import pandas as pd
from itertools import product
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys 

class mylog:
    @staticmethod
    def info(*args, **kwargs):
        print(*args, **kwargs)

path = "/Users/wangyangzuo/Desktop/公司/sd_forward.json"

total_graph = json.load(open(path,'r'))
graph         = total_graph['graph']
links         = total_graph['links']
global_idx    = 10000000
reverse_links = defaultdict(list)
in_degree     = defaultdict(int)
id_info       = dict()
id_name       = dict()
is_fuse       = int(sys.argv[1])
output_file_name = sys.argv[2]


def walk_for_moudle():
    next_nodes = [[graph, None]]
    param = 0
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        if not cur_node[k]["children"]:
            continue
        if len(cur_node[k]["children"]) == 1:
            param += calc_module_paraments(cur_node[k])
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return param

param = walk_for_moudle()

def build_reverse_links(links):
    global reverse_links, in_degree
    reverse_links = defaultdict(list)
    in_degree     = defaultdict(int)
    for k,v in links.items():
        for link in v:
            reverse_links[link].append(k)
            in_degree[k] += 1

build_reverse_links(links)

def build_id_info():
    global id_info, id_name
    next_nodes  = [[graph, None]]
    id_info = dict()
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id       = list(node.keys())[0]
        if "children" not in node[key_id] or not node[key_id]['children']:
            id_info[key_id] = node[key_id]
            continue
        for child in node[key_id]['children']:
            next_nodes.append([child, node])
    id_name = {i:id_info[i]["name"] for i in id_info.keys()}

build_id_info()

def bfs_search(start_id, names, preset=None):
    t_s = [start_id]
    vis = set().union().union(preset)
    ends = [id_name[str(start_id)]] + names
    res = []
    while t_s:
        t = t_s.pop(0)
        if id_name[str(t)] != ends[len(res)]:
            return []
        if id_name[str(t)] == ends[len(res)]:
            res.append(t)
        vis.add(t)
        if len(res) == len(ends):
            return res
        if str(t) in links:
            for link in links[str(t)]:
                if link not in vis:
                    t_s.append(int(link))
    return res

def bfs_search_bwd(end_id, names, preset=None):
    t_s = [end_id]
    vis = set().union(preset)
    ends = [id_name[str(end_id)]] + names
    res = []
    while t_s:
        t = t_s.pop(0)
        if id_name[str(t)] != ends[len(res)]:
            return []
        if id_name[str(t)] == ends[len(res)]:
            res.append(t)
        vis.add(t)
        if len(res) == len(ends):
            return res
        if int(t) in reverse_links:
            for link in reverse_links[int(t)]:
                if link not in vis:
                    t_s.append(link)
    return res

def help_parse_pattern(pattern:str):
    assert "(" not in pattern
    top_down = True if "->" in pattern else False
    pattern = pattern.split("<-") if not top_down else pattern.split("->")
    return pattern, top_down

def parse_pattern(pattern:str):
    # to->linear->linear->to->(mul<-linear)->add 
    # 返回 [-1,-1, [to, linear, linear, to, mul, linear, add], top_down], [0,4 [mul, linear], down_top]
    res = []
    content = re.findall(r'\(.*?\)', pattern)
    if not content:
        mylog.info("simple pattern")
        p, top_down = help_parse_pattern(pattern)
        return [[-1,-1,p,top_down]]
    for idx, c in enumerate(content):
        pattern = pattern.replace(c, "*")
    fp, ftop_down = help_parse_pattern(pattern)
    for idx, c in enumerate(content):
        idx = fp.index("*", idx)
        p, top_down = help_parse_pattern(c[1:-1])
        mylog.info(p, fp, idx)
        fp[idx] = p[0]
        res.append([0, idx, p, top_down])
    return [[-1,-1,fp,ftop_down]] + res

def search_cur_patten(graph, links, pattern):
    pre, node_id, p, top_down = pattern
    start_name = p[0]
    the_other  = p[1:]
    next_nodes  = [[graph, None]]
    pre_set = set() if pre == -1 else set(pre)
    while next_nodes:
        node, parent = next_nodes.pop(0)
        key_id       = list(node.keys())[0]
        if (node_id == -1 and node[key_id]['name'] == start_name) or int(node_id) == int(key_id):
            res = bfs_search_bwd(int(key_id), the_other, pre_set) if not top_down else bfs_search(int(key_id), the_other, pre_set)
            if not res: continue
            return res
        if "children" not in node[key_id] or not node[key_id]['children']:
            continue
        for child in node[key_id]['children']:
            next_nodes.append([child, node])

def search_pattern(graph, links, raw_pattern):
    pattern = parse_pattern(raw_pattern)
    res = search_cur_patten(graph, links, pattern[0])
    if not res: return res
    for p in pattern[1:]:
        p[0] = res
        idx = p[1]
        p[1] = res[idx]
        tmp = search_cur_patten(graph, links, p)
        for i in tmp:
            if i not in res:
                res.append(i)
    return res

def build_a_fuse_op(name, depth=None, input_shape=None, output_shape=None, inputs_dtypes=None, outputs_dtypes=None,  comment=""):
    global global_idx
    idx = global_idx
    global_idx += 1
    t = {
        "name": name,
        "depth": depth,
        "idx": idx,
        "input_shape"  : input_shape,
        "output_shape" : output_shape,
        "input_dtype"  : inputs_dtypes,
        "output_dtype" : outputs_dtypes,
        "comment" : comment,
        "children": None
    }
    return {str(idx): t}

def handle_link(cur_id:int|str, replace_id=None):
    cur_id = int(cur_id)
    build_reverse_links(links)
    father_ids = reverse_links[cur_id]
    # import pdb;pdb.set_trace()
    # mylog.info(f"cur_id, {cur_id}, cur_name, {id_name[str(cur_id)]}, father_ids, {father_ids}, replace_id, {replace_id}")
    if replace_id:
        for father_id in father_ids:
            idx = links[father_id].index(cur_id)
            links[father_id][idx] = int(replace_id)
        if str(replace_id) not in links or not links[str(replace_id)]:
            links[str(replace_id)] = links[str(cur_id)]
        del links[str(cur_id)]
    else:
        for father_id in father_ids:
            idx = links[father_id].index(cur_id)
            links[father_id].pop(idx)
        del links[str(cur_id)]
    build_id_info()

def build_attention_op(badbmm_op):
    depth = badbmm_op['depth']
    shape = badbmm_op['input_shape']
    input_shape  = [ shape[1], shape[2] ,shape[2]]
    output_shape = [ shape[1]]
    input_dtype  = badbmm_op['input_dtype']
    output_dtype = badbmm_op['output_dtype']
    comment      = "fuse softmax(qk/d)*v op with memory efficient"
    return build_a_fuse_op("scaled_dot_attention", depth, input_shape, output_shape, input_dtype, output_dtype, comment)

def attention_match_and_rewrite(graph, links, raw_pattern="baddbmm->softmax->to->bmm"):
    res         = search_pattern(graph, links, raw_pattern)
    if not res: return graph, links, False
    print(res, [ id_name[str(i)] for i in res])
    next_nodes  = [[graph, None]]
    attenion_op = None
    flag        = False
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id       = list(node.keys())[0]
        parent_id    = list(parent.keys())[0] if parent else None
        if int(key_id) == int(res[0]):
            flag = True
            attenion_op = build_attention_op(node[key_id])
            idx         = parent[parent_id]['children'].index(node)
            parent[parent_id]['children'][idx] = attenion_op
            attention_op_id = list(attenion_op.keys())[0]
            handle_link(key_id, attention_op_id)
        elif int(key_id) == int(res[-1]):
            parent[parent_id]['children'].remove(node)
            handle_link(key_id, attention_op_id)
        elif int(key_id) in res:
            parent[parent_id]['children'].remove(node)
            handle_link(key_id)
        if "children" not in node[key_id] or not node[key_id]['children']:
            continue
        for child in node[key_id]['children'][::-1]:
            next_nodes.append([child, node])
    return graph, links, flag

def build_fuse_lora_op(node_id_lists):
    to           = id_info[ str(node_id_lists[0]) ]
    linear1      = id_info[ str(node_id_lists[1]) ]
    linear2      = id_info[ str(node_id_lists[2]) ]
    linear3      = id_info[ str(node_id_lists[-1]) ]
    weights      = [linear3["input_shape"][0][2], linear3["output_shape"][0][2]]
    lora_rank    = linear1["output_shape"][0][2]
    name         = "fuse_lora"
    depth        = linear1['depth']
    input_shape  = [ linear1['input_shape'][0] ]
    output_shape = linear2['output_shape']
    input_dtype  = [ to["input_dtype"][0] ]
    output_dtype = [ to["input_dtype"][0] ]
    comment      = f"kernel_shape=[{weights[0]},{weights[1]}], rank={lora_rank}, weight_dtype=float16, lora_dtype=float32, fuse to->linear->linear->to->mul->(add<-linear) op"
    return build_a_fuse_op(name, depth, input_shape, output_shape, input_dtype, output_dtype, comment)

def lora_linear_match_and_rewrite(graph, links, raw_pattern="to->linear->linear->to->mul->(add<-linear)"):
    res = search_pattern(graph, links, raw_pattern)
    if not res: return graph, links, False
    res = [int(i) for i in res]
    mylog.info(res, [ id_name[str(i)] for i in res])
    next_nodes = [[graph, None]]
    flag = False
    fuse_lora_op = None
    fuse_lora_op_id = None
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id = list(node.keys())[0]
        parent_id = list(parent.keys())[0] if parent else None
        if int(key_id) == int(res[0]):
            flag         = True
            idx          = parent[parent_id]['children'].index(node)
            parent[parent_id]['children'][idx] = fuse_lora_op
            parent[parent_id]['need_train'] = True
            fuse_lora_op_id = list(fuse_lora_op.keys())[0]
            handle_link(key_id, fuse_lora_op_id)
            continue
        elif int(key_id) == int(res[-1]):
            fuse_lora_op = build_fuse_lora_op(res)
            parent[parent_id]['children'].remove(node)
            handle_link(key_id)
            continue
        elif int(key_id) == int(res[-2]):
            parent[parent_id]['children'].remove(node)
            handle_link(key_id, fuse_lora_op_id)
            continue
        elif int(key_id) in res:
            parent[parent_id]['children'].remove(node)
            handle_link(key_id)
            continue
        if "children" not in node[key_id] or not node[key_id]['children']:
            continue
        for child in node[key_id]['children'][::-1]:
            next_nodes.append([child, node])
    return graph, links, flag

def remove_attention_redundancy_nodes(graph, links, pattern="reshape->permute->reshape->transpose->baddbmm"):
    res = search_pattern(graph, links, pattern)
    # import pdb;pdb.set_trace()
    if not res: return graph, links, False
    mylog.info(res, [ id_name[str(i)] for i in res])
    next_nodes = [[graph, None]]
    flag = False
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id = list(node.keys())[0]
        parent_id = list(parent.keys())[0] if parent else None
        if int(key_id) == res[0]:
            # import pdb;pdb.set_trace()
            parent[parent_id]['children'].remove(node)
            handle_link(key_id, res[-1] )
            continue
        elif int(key_id) in res[:-1]:
            parent[parent_id]['children'].remove(node)
            handle_link(key_id)
            continue
        elif int(key_id) == res[-1]:
            flag = True
            break
        if "children" not in node[key_id] or not node[key_id]['children']:
            continue
        for child in node[key_id]['children'][::-1]:
            next_nodes.append([child, node])
    return graph, links, flag

def remove_attention_redundancy_nodes2(graph, links, pattern="reshape->permute->reshape->baddbmm"):
    return remove_attention_redundancy_nodes(graph, links, pattern)

def remove_attention_redundancy_nodes3(graph, links, pattern="reshape->permute->reshape->bmm"):
    return remove_attention_redundancy_nodes(graph, links, pattern)

fuse_pattern = [
    lora_linear_match_and_rewrite,
    # attention_match_and_rewrite
    # remove_attention_redundancy_nodes,
    # remove_attention_redundancy_nodes2,
    # remove_attention_redundancy_nodes3,
]

def do_fuse():
    global graph, links
    while 1:
        flag = False
        for pattern in fuse_pattern:
            graph, links, f = pattern(graph, links)
            flag |= f
        if not flag:
            break

fuse_pattern2 = [
    attention_match_and_rewrite
]

def do_fuse2():
    global graph, links
    while 1:
        flag = False
        for pattern in fuse_pattern2:
            graph, links, f = pattern(graph, links)
            flag |= f
        if not flag:
            break

if is_fuse:
    do_fuse()
    do_fuse2()


def remove_empty_module():
    global graph
    next_nodes = [[graph, None]]
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id = list(node.keys())[0]
        parent_key = list(parent.keys())[0] if parent else None
        if "children" not in node[key_id] or node[key_id]['children'] == None:
            continue
        if len(node[key_id]['children']) == 0:
            parent[parent_key]['children'].remove(node)
        for child in node[key_id]['children'][::-1]:
            next_nodes.append([child, node])

remove_empty_module()


# find a transformer 
# from layernorm to layernorm
# def find_transformer():
#     next_nodes = [[graph, None]]
#     transformer_nodes = []
#     more_transformer = []
#     flag = False
#     while next_nodes:
#         node, parent = next_nodes.pop()
#         key_id = list(node.keys())[0]
#         if node[key_id]['name'] == "conv2d":
#             if flag:
#                 more_transformer.append(transformer_nodes)
#                 transformer_nodes = []
#             else:
#                 flag = True
#         if "children" not in node[key_id] or node[key_id]['children'] == None:
#             if node[key_id]['name'] == "conv2d":
#                 flag = False
#                 transformer_nodes = []
#             if flag:
#                 transformer_nodes.append(key_id)
#             continue
#         for child in node[key_id]['children'][::-1]:
#             next_nodes.append([child, node])
#     return more_transformer

def find_transformer():
    next_nodes = [[graph, None]]
    res = []
    single_basic_transformer = []
    flag = False
    while next_nodes:
        node, parent = next_nodes.pop()
        key_id = list(node.keys())[0]
        if node[key_id]['name'] == "BasicTransformerBlock":
            flag = True
        if node[key_id]["name"] == "Transformer2DModel":
            if flag:
                res.append(single_basic_transformer)
                single_basic_transformer = []
                flag = False
            flag = False
        if "children" not in node[key_id] or node[key_id]['children'] == None:
            if flag:
                single_basic_transformer.append(key_id)
            continue
        for child in node[key_id]['children'][::-1]:
            next_nodes.append([child, node])
    return res

do_select = True
need_calc_nodes = find_transformer()[0] # debug 
# with open("graph.json", "w") as f:
#     json.dump(graph, f, indent=4)
# import pdb;pdb.set_trace()


def find_grad_nodes(graph):
    grad_node_ids   = []
    grad_module_ids = []
    next_nodes = [[graph, None]]
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        if father_k and father_node[father_k]['need_train']:
            grad_node_ids.append(k)
            grad_module_ids.append([father_k,father_node[father_k]["name"], father_node[father_k]["input_dtype"][0],father_node[father_k]["comment"]])
        if not cur_node[k]["children"]:
            continue
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return grad_node_ids, grad_module_ids

mylog.info(">>>>> find grad nodes")

grad_node_ids,grad_module = find_grad_nodes(graph)

def get_linear_grad_info(single_module):
    dtype = single_module[2]
    info  = single_module[3]
    in_features  = int(info.split("in_features=")[1].split(",")[0])
    out_features = int(info.split("out_features=")[1].split(",")[0])
    bias = info.split("bias=")[1].split(",")[0] == "True"  
    mem_info = in_features * out_features * dtype_map[dtype] + (out_features if bias else 0) * dtype_map[dtype]
    return mem_info, [out_features, in_features], dtype

def parse_grad_module_info(grad_module):
    grads = []
    grad_mem = 0
    adam_mem = 0
    for each_grad in grad_module:
        if each_grad[1] == "Linear":
            mem_info, shape, dtype = get_linear_grad_info(each_grad)
            grad_mem += mem_info
            adam_mem += mem_info * 2 if each_grad[2] == "float32" else mem_info * 4
            grads.append([shape, dtype])
        # grads.append(each_grad)
    return grads, grad_mem, adam_mem

grads, grad_mem, adam_mem = parse_grad_module_info(grad_module)

mylog.info(">>>>> find all activation node")
starts = [i for i in grad_node_ids]
activation_nodes = set()
while starts:
    start = starts.pop()
    activation_nodes.add(start)
    if start not in links:
        continue
    for link in links[start]:
        if str(link) not in activation_nodes:
            starts.append(str(link))

mylog.info(">>>>> find all loss nodes")
# find all loss nodes
starts = [i for i in grad_node_ids]
loss_nodes =[]
vis = set()
while starts:
    start = starts.pop()
    vis.add(start)
    if str(start) not in links:
        loss_nodes.append(start)
        continue
    for link in links[str(start)]:
        if link not in vis:
            starts.append(link)

mylog.info(">>>>> match shape")

model_param = total_graph.get('model_param', {
    "n"     : 1,
    "c"     : 4,
    "h"     : 64,
    "w"     : 64,
    "seq"   : 77,
    "dtype" : "float16",
    "rank"  : 4
})

hw_region  = [64*64, 64*64/4, 64*64/16, 64*64/64]
h_w_region = [64, 32, 16, 8]

class layout:
    # 4 dim
    n_model_h_w       = "n_model_h_w"
    n_h_w_model       = "n_h_w_model"
    n_model_hw_model  = "n_model_hw_model"
    n_hw_model_model  = "n_hw_model_model"
    n_seq_model_model = "n_seq_model_model"
    n_model_model_model = "n_model_model_model"
    n_model_seq_model = "n_model_seq_model"
    # 3 dim
    n_seq_model       = "n_seq_model"
    n_seq_r           = "n_seq_r"
    nmodel_model_hw   = "nmodel_model_hw"
    nmodel_hw_seq     = "nmodel_hw_seq"
    nmodel_model_seq  = "nmodel_model_seq"
    nmodel_seq_model  = "nmodel_seq_model"
    n_hw_r            = "n_hw_r"
    # 3 dim
    n_hw_model        = "n_hw_model"
    n_hw_hw           = "n_hw_hw"
    nmodel_hw_hw      = "nmodel_hw_hw"
    nmodel_hw_model   = "nmodel_hw_model"
    # 2 dim
    n_model           = "n_model"
    # 1 dim
    n                 = "n"
    model             = "model"
 
    @staticmethod
    def match_shape(shape: list):
        if len(shape) == 1:
            if shape[0] == model_param['n']:
                return layout.n, [1]
            return layout.model, [0]
        if len(shape) == 2:
            return layout.n_model, [1,0]
        if len(shape) == 3:
            if shape[0] == model_param['n']:
                if shape[1] == model_param["seq"]:
                    if shape[2] == model_param["rank"]:
                        return layout.n_seq_r, [1,shape[1] / model_param["seq"], shape[2] / model_param["rank"]]
                    else:
                        return layout.n_seq_model, [1, shape[1] / model_param["seq"], 0]
                elif shape[2] == model_param["rank"]:
                    return layout.n_hw_r, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / model_param["rank"]]
                elif shape[1] == shape[2]:
                    return layout.n_hw_hw, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / (model_param["h"] * model_param["w"])]
                else:
                    return layout.n_hw_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0]
            # 3 dim not starts with n
            elif shape[1] == model_param["seq"]:
                return layout.nmodel_seq_model, [1, shape[1] / model_param["seq"], 0]
            elif shape[2] == model_param["seq"] and shape[1] in h_w_region:
                # nmodel_hw_seq
                return layout.nmodel_hw_seq, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / model_param["seq"]]
            elif shape[2] == model_param["seq"]:
                # nmodel_model_seq
                return layout.nmodel_model_seq, [1, 0, shape[2] / model_param["seq"]]
            elif shape[1] == shape[2] and shape[1] in hw_region:
                # nmodel_hw_hw
                return layout.nmodel_hw_hw, [1, shape[1] / (model_param["h"] * model_param["w"]), shape[2] / (model_param["h"] * model_param["w"])]
            elif shape[2] in hw_region:
                return layout.nmodel_model_hw, [1, 0, shape[2] / (model_param["h"] * model_param["w"])]
            else:
                return layout.nmodel_hw_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0]
        if len(shape) == 4:
            if shape[0] == model_param['n']:
                if shape[1] == model_param["seq"]:
                    return layout.n_seq_model_model, [1, shape[1] / model_param["seq"], 0, 0]
                elif shape[1] in h_w_region and shape[2] in h_w_region and shape[1] == shape[2]:
                    return layout.n_h_w_model, [1, shape[1] / model_param["h"], shape[2] / model_param["w"], 0]
                elif shape[2] in h_w_region and shape[3] in h_w_region:
                    return layout.n_model_h_w, [1, 0, shape[2] / model_param["h"], shape[3] / model_param["w"]]
                elif shape[1] in hw_region:
                    return layout.n_hw_model_model, [1, shape[1] / (model_param["h"] * model_param["w"]), 0, 0]
                elif shape[2] in hw_region:
                    return layout.n_model_hw_model, [1, 0, shape[2] / (model_param["h"] * model_param["w"]), 0]
                elif shape[2] == model_param["seq"]:
                    return layout.n_model_seq_model, [1, 0, shape[2] / model_param["seq"], 0]
                else:
                    return layout.n_model_model_model, [1, 0, 0, 0]
            else:
                import pdb;pdb.set_trace()
                print("shape error")
        return 0,[0]

    @staticmethod
    def match_layout(node):
        input_shapes = node["input_shape"]
        output_shapes = node["output_shape"]
        node["input_shape_layout"]  = [layout.match_shape(i) for i in input_shapes]
        node["output_shape_layout"] = [layout.match_shape(i) for i in output_shapes]
        return node

    @staticmethod
    def match_layout_source_node(node):
        k = list(node.keys())[0]
        input_shapes = node[k]["input_shape"]
        output_shapes = node[k]["output_shape"]
        node[k]["input_shape_layout"]  = [layout.match_shape(i) for i in input_shapes]
        node[k]["output_shape_layout"] = [layout.match_shape(i) for i in output_shapes]
        return node

    @staticmethod
    def fix_single_shape(shape, cur_layout, rate, param):
        # param 修复过的 only support nhw
        if cur_layout == layout.n:
            return [param["n"]]
        if cur_layout == layout.model:
            return shape
        if cur_layout == layout.n_model:
            return [param["n"], shape[1]]
        if cur_layout == layout.n_seq_r:
            return [param["n"], shape[1], shape[2]]
        if cur_layout == layout.n_seq_model:
            return [param["n"], shape[1], shape[2]]
        if cur_layout == layout.n_hw_r:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.n_hw_hw:
            return [param["n"], param["h"] * param["w"] * rate[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.n_hw_model:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_seq_model:
            return [shape[0] * param["n"], shape[1], shape[2]]
        if cur_layout == layout.nmodel_hw_seq:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_model_seq:
            return [shape[0] * param["n"], shape[1], shape[2]]
        if cur_layout == layout.nmodel_hw_hw:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.nmodel_hw_model:
            return [shape[0] * param["n"], param["h"] * param["w"] * rate[1], shape[2]]
        if cur_layout == layout.nmodel_model_hw:
            return [shape[0] * param["n"], shape[1], param["h"] * param["w"] * rate[2]]
        if cur_layout == layout.n_seq_model_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        if cur_layout == layout.n_h_w_model:
            return [param["n"], param["h"] * rate[1], param["w"] * rate[2], shape[3]]
        if cur_layout == layout.n_model_h_w:
            return [param["n"], shape[1], param["h"] * rate[2], param["w"] * rate[3]]
        if cur_layout == layout.n_model_model_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        if cur_layout == layout.n_hw_model_model:
            return [param["n"], param["h"] * param["w"] * rate[1], shape[2], shape[3]]
        if cur_layout == layout.n_model_hw_model:
            return [param["n"], shape[1], param["h"] * param["w"] * rate[2], shape[3]]
        if cur_layout == layout.n_model_seq_model:
            return [param["n"], shape[1], shape[2], shape[3]]
        print("error: not support layout")
        import pdb;pdb.set_trace()
    
    @staticmethod
    def fix_shape(node, param):
        # param: {n,h,w,r,seq}
        k                = list(node.keys())[0]
        input_shapes     = node[k]["input_shape"]
        output_shapes    = node[k]["output_shape"]
        input_layer_out  = node[k]["input_shape_layout"]
        output_layer_out = node[k]["output_shape_layout"]
        new_input_shape  = []
        new_output_shape = []
        # print("fix shape", input_shapes, output_shapes)
        # print("layout ", input_layer_out, output_layer_out)
        for i in range(len(input_shapes)):
            cur_layout, rate = input_layer_out[i]
            t = layout.fix_single_shape(input_shapes[i], cur_layout, rate, param)
            new_input_shape.append([int(i) for i in t])
        for i in range(len(output_shapes)):
            cur_layout, rate = output_layer_out[i]
            t = layout.fix_single_shape(output_shapes[i], cur_layout, rate, param)
            new_output_shape.append([int(i) for i in t])
        # print("after",  new_input_shape, new_output_shape)
        # print(">>>>>>>>>>>>> ", node[k]["name"], k, " <<<<<<<<<<<<<<<<")
        node[k]["input_shape"] = new_input_shape
        node[k]["output_shape"]= new_output_shape
        return node

# import pdb;pdb.set_trace()
# init: 输入的权重，梯度权重 adam 
weights         = total_graph.get('weights', 2041.164e6 )
grad_weights    = grad_mem
adam_weights    = adam_mem

all_grad_shapes = grads
first_key = list(graph.keys())[0]
start_depth = 0

def build_preview_mem():
    
    pass


def match_shape():
    next_nodes = [[graph, None]]
    vis = set()
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        if k in vis: continue
        if not cur_node[k]["children"]:
            layout.match_layout_source_node(cur_node)
            vis.add(k)
            continue
        else:
            layout.match_layout_source_node(cur_node)
        vis.add(k)
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
mylog.info(">>>>> match shape")
match_shape()
source_graph = deepcopy(graph)



# fix shape 


def fix_shape_total_graph(cur_shapes):
    next_nodes = [[graph, None]]
    vis = set()
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        if k in vis: continue
        if not cur_node[k]["children"]:
            layout.fix_shape(cur_node, cur_shapes)
            continue
        else:
            layout.fix_shape(cur_node, cur_shapes)
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])

def get_output_mem(node):
    k = list(node.keys())[0]
    name = node[k]['name']
    if name in ["reshape", "__getitem__", "to","chunk", "contiguous", "float", "permute", "transpose"]:
        return 0
    output_shapes = node[k]['output_shape']
    output_dtype = node[k]['output_dtype']
    output_dtypes = [ dtype_map[i] for i in output_dtype]
    out_mem = sum([np.prod(i) * j for i, j in zip(output_shapes, output_dtypes)])
    return int(out_mem)

def get_output_mem_node(node):
    k = node["id"]
    name = node['name']
    if name in ["reshape", "__getitem__", "to","chunk", "contiguous", "float", "permute", "transpose"]:
        return 0
    output_shapes = node['output_shape']
    output_dtype = node['output_dtype']
    output_dtypes = [ dtype_map[i] for i in output_dtype]
    out_mem = sum([np.prod(i) * j for i, j in zip(output_shapes, output_dtypes)])
    return int(out_mem)

def activation_has_outputs_check_node(node):
    k = node["id"]
    name = node["name"]
    if name in ["baddbmm", "mul", "add", "cat"]:
        return False
    if k in activation_nodes:
        return True
    return False

def activation_has_outputs_node(node):
    k = list(node.keys())[0]
    name = node[k]['name']
    if name in ["baddbmm", "mul", "add"]:
        return False
    if k in activation_nodes:
        return True
    return False

def walk_for_training_mem_activation_mem(graph):
    # grad_node_ids = []
    next_nodes = [[graph, None]]
    module_stack = []
    module_stack_node = []
    idx = 0
    activation_mem_usage = defaultdict(int)
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        while father_k and module_stack and father_k != module_stack[-1]:
            t_k = module_stack.pop()
            t   = module_stack_node.pop()
            cur_t_k = module_stack[-1] if module_stack else None
            cur_t   = module_stack_node[-1] if module_stack else None
            if cur_t_k:
                cur_t[cur_t_k]["module_mem"] += get_output_mem(t)
                cur_t[cur_t_k]["graph_mem"]  += t[t_k]["graph_mem"]
                cur_t[cur_t_k]["cur_mem"] = max(cur_t[cur_t_k]["module_mem"], cur_t[cur_t_k]["graph_mem"] )
                t[t_k]["max_mem"]         = max(t[t_k]["max_mem"], t[t_k]["module_mem"])
                cur_t[cur_t_k]["max_mem"] = max(cur_t[cur_t_k]["max_mem"], t[t_k]["max_mem"], cur_t[cur_t_k]["module_mem"] )
        
        if cur_node[k]["children"]:
            module_stack.append(k)
            module_stack_node.append(cur_node)
            cur_node[k]["module_mem"] = 0
            cur_node[k]["graph_mem"]  = 0
            cur_node[k]["max_mem"]    = 0
            cur_node[k]["cur_mem"]    = max(father_node[father_k]["module_mem"], father_node[father_k]["graph_mem"] )if father_k else 0
        
        if not cur_node[k]['children']:
            cur_node[k]["idx"] = idx
            idx += 1
            cur_mem = get_output_mem(cur_node)
            cur_node[k]['cur_mem'] = father_node[father_k]["cur_mem"] + cur_mem
            cur_node[k]["module_mem"] = cur_mem

            if activation_has_outputs_node(cur_node):
                cur_node[k]["graph_mem"] = cur_mem

            father_node[father_k]["module_mem"] += cur_mem
            father_node[father_k]["cur_mem"]    += cur_mem
            if activation_has_outputs_node(cur_node):
                activation_mem_usage[cur_node[k]["name"]] += cur_mem
                father_node[father_k]["graph_mem"] += cur_mem
            continue

        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return activation_mem_usage

def forward_dma_ops_calc():
    res_set = {}
    def handle_record_node(node, pre):
        node_key                 = list(node.keys())[0]
        pre_key                  = list(pre.keys())[0]
        res_node                 = {}
        depth                    = node[node_key]['depth']
        node[node_key]['info']   = pre[pre_key]["comment"]
        if node[node_key]["name"] == "empty-pass":
            return 
        if node[node_key]["name"] in fix_op:
            node[node_key] = fix_op[node[node_key]["name"]](node[node_key])
        # layout.match_layout(node[node_key])
        node[node_key]                 = ops_info[node[node_key]["name"]](node[node_key])
        res_node["name"]               = node[node_key]["name"]
        res_node['id']                 = node_key
        res_node['depth']              = node[node_key]['depth']
        node[node_key]['info']         = pre[pre_key]["comment"][:200] if depth > 1 else ""
        res_node['path']               = node[node_key]['path']
        res_node["info"]               = pre[pre_key]["comment"][:200] if depth > 1 else ""
        if res_node["name"] == "fuse_lora":
            res_node["info"] = node[node_key]["comment"]
        res_node['align_input_shape']  = node_fn[res_node['name']](node)
        res_node["input_shape"]        = node[node_key]['input_shape']
        res_node["output_shape"]       = node[node_key]["output_shape"]
        res_node['align_output_shape'] = [ handle_input_shape(node[node_key]['output_shape'][0])]
        res_node['align_input_dtype']  = node[node_key]['input_dtype'][0]
        res_node['align_output_dtype'] = node[node_key]['output_dtype'][0]
        res_node['input_dtype']        = node[node_key]['input_dtype']
        res_node['output_dtype']       = node[node_key]['output_dtype']
        res_node["ops"]                = node[node_key]["ops"]
        res_node["vector_ops"]         = node[node_key]["vector_ops"]
        res_node["cube_ops"]           = node[node_key]["cube_ops"]
        res_node["s2l_dma"]            = node[node_key]["s2l_dma"]
        res_node["l2s_dma"]            = node[node_key]["l2s_dma"]
        res_node["input_shape_layout"] = node[node_key]["input_shape_layout"]
        res_node["output_shape_layout"] = node[node_key]["output_shape_layout"]
        res_node["cur_mem"]      = node[node_key]["cur_mem"]
        res_node["forward_dma"]  = res_node["s2l_dma"] + res_node["l2s_dma"]
        if res_node["name"] == "conv2d":
            res_node["kernel_shape"] = res_node["align_input_shape"][1]
        # res_set.append(res_node)
        res_set[node_key] = res_node
        return res_node
    next_nodes = [[graph, None]]
    # need_mem = defaultdict(dict)
    # tables = []
    while next_nodes:
        cur_node, father_node = next_nodes.pop()
        k = list(cur_node.keys())[0]
        father_k = list(father_node.keys())[0] if father_node else None
        if cur_node[k]['depth'] >= 0:
            if "path" not in father_node[father_k]:
                cur_node[k]["path"] = [ cur_node[k]["name"] ]
            else:
                cur_node[k]["path"] = father_node[father_k]["path"] + [ cur_node[k]["name"] ]
        if not cur_node[k]["children"]:
            handle_record_node(cur_node, father_node)
            continue
        for child in cur_node[k]['children'][::-1]:
            next_nodes.append([child, cur_node])
    return res_set

# get back_grad_idx
def insert_back_grad_idx(node):
    node_id = node['id']
    if node_id not in activation_nodes: return node
    name = node['name']
    if node_id in grad_node_ids:
        node["leaf"] = True
    else:
        node["leaf"] = False
    if node["leaf"]:
        if name == "conv":
            # check bais 
            has_bias = node.get("bias", False)
            node["back_grad_idx"] = list(range(3)) if has_bias else list(range(2))
        elif name == "linear":
            # check bias
            has_bias = node.get("bias", False)
            node["back_grad_idx"] = list(range(3)) if has_bias else list(range(2))
        elif name == "layer_norm":
            # check affine
            node["back_grad_idx"] = list(range(3))
        elif name == "group_norm":
            # check affine
            node["back_grad_idx"] = list(range(3))
        elif name == "fuse_lora":
            node["back_grad_idx"] = list(range(3))
        else:
            print("error: not support leaf node")
            import pdb;pdb.set_trace()
            exit(1)
    else:
        # 看输入是否是激活节点，如果是就算
        # 有一些只有一个输入或者是输入较少，比如说mul 构造隐藏的tensor节点
        input_ids = [i for i in reverse_links[int(node_id)] if i in activation_nodes]
        # match inputs node
        input_num = len(node["input_shape"])
        if name in ["mul", "add", "div", "sub"] and input_num == 1:
            node["back_grad_idx"] = [0]
        if input_num == len(input_ids):
            node["back_grad_idx"] = list(range(input_num))
        else:
            input_ids_shapes = [res_set[i]["output_shape"][0] for i in input_ids]
            res = []
            for i in range(input_num):
                if node["input_shape"][i] in input_ids_shapes:
                    res.append(i)
            node["back_grad_idx"] = res
    return node

def prepare_bwd_node_outputs_calc(res_set):
    for k in res_set.keys():
        res_set[k] = insert_back_grad_idx(res_set[k])
    return res_set

def calc_bwd_bdc_dma(res_set):
    build_reverse_links(links)
    backward_start = [ str(i) for i in loss_nodes]
    print("backward start",backward_start)
    grad_set = set(grad_node_ids)
    cur_back_in_degress = deepcopy(in_degree)
    while backward_start:
        start = backward_start.pop(0)
        if start in activation_nodes and start in res_set:
            # calculate 
            # backward_node_fn
            node = res_set[start]
            name = node["name"]
            calc_dma_backops(node)
            node = backward_node_fn[name](node)
            # remove mem usage
            if activation_has_outputs_check_node(node):
                node["back_mem"] = node["cur_mem"] - get_output_mem_node(node)
        if start in grad_set:
            grad_set.remove(start)
        if len(grad_set) == 0:
            print("have found all the backward nodes")
            break
        for v in reverse_links[int(start)]:
            cur_back_in_degress[v] -= 1
            if cur_back_in_degress[v] == 0:
                backward_start.append(v)
    return res_set

def calc_bwd_grad_chain(res_set):
    build_reverse_links(links)
    backward_start = [ str(i) for i in loss_nodes]
    print("backward start",backward_start)
    grad_set = set(grad_node_ids)
    cur_back_in_degress = deepcopy(in_degree)
    grad_add_count = 0
    while backward_start:
        start = backward_start.pop(0)
        if start in activation_nodes and start in res_set:
            # calculate 
            # backward_node_fn
            node = res_set[start]
            name = node["name"]
            grad_in    = len([i for i in links[node['id']] if str(i) in activation_nodes])
            output_len = len(node['output_shape'])
            if grad_in > output_len:
                if output_len > 1:
                    print("error: not support")
                    import pdb;pdb.set_trace()
                # grad will add by auto grad chain
                node["grad_chain_count"] = grad_in - output_len
                grad_add_count += grad_in - output_len
                node["grad_chain_dma"]   = np.prod(node["output_shape"][0]) * dtype_map[node["output_dtype"][0]] * 3 * (grad_in - output_len)
                node["grad_chain_ops"]   = np.prod(node["output_shape"][0]) * (grad_in - output_len)
        if start in grad_set:
            grad_set.remove(start)
        if len(grad_set) == 0:
            print("have found all the backward nodes")
            break
        for v in reverse_links[int(start)]:
            cur_back_in_degress[v] -= 1
            if cur_back_in_degress[v] == 0:
                backward_start.append(v)
    print("grad add count", grad_add_count)
    return res_set

def handle_tiu_time_calc(res_set,k):
    rate = params["1684x f32 tiu"] / params["1684x f16 tiu"]
    res_set[k]["1684x_forward_tiu_vec_time"]  = (res_set[k]["vector_ops"] / params["1684x vector tiu"]) * 1 if res_set[k]["input_dtype"][0] == "float16" else rate 
    res_set[k]["1684x_forward_tiu_cube_time"] = (res_set[k]["cube_ops"] / params["1684x cuba tiu"]) * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["1684x_forward_ops_time"]      = res_set[k]["1684x_forward_tiu_vec_time"] + res_set[k]["1684x_forward_tiu_cube_time"]
    res_set[k]["2260_forward_tiu_vec_time"]   = (res_set[k]["vector_ops"] / params["2260 vector tiu"]) * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["2260_forward_tiu_cube_time"]  = (res_set[k]["cube_ops"] / params["2260 cuba tiu"]) * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["2260_forward_ops_time"]       = res_set[k]["2260_forward_tiu_vec_time"] + res_set[k]["2260_forward_tiu_cube_time"]

def handle_bwd_tiu_time_calc(res_set,k):
    rate = params["1684x f32 tiu"] / params["1684x f16 tiu"]
    res_set[k]["1684x_backward_tiu_vec_time"]  = (res_set[k]["back_vector_ops"] / params["1684x vector tiu"])   * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["1684x_backward_tiu_cube_time"] = (res_set[k]["back_cube_ops"] / params["1684x cuba tiu"])       * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["1684x_backward_ops_time"]      = res_set[k]["1684x_backward_tiu_vec_time"] + res_set[k]["1684x_backward_tiu_cube_time"]
    res_set[k]["2260_backward_tiu_vec_time"]   = (res_set[k]["back_vector_ops"] / params["2260 vector tiu"])    * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["2260_backward_tiu_cube_time"]  = (res_set[k]["back_cube_ops"] / params["2260 cuba tiu"])        * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["2260_backward_ops_time"]       = res_set[k]["2260_backward_tiu_vec_time"] + res_set[k]["2260_backward_tiu_cube_time"]

def handle_grad_add_time_calc(res_set,k):
    rate = params["1684x f32 tiu"] / params["1684x f16 tiu"]
    res_set[k]["grad_add_1684x_backward_tiu_vec_time"]  = (res_set[k]["grad_chain_ops"] / params["1684x vector tiu"])   * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["grad_add_1684x_backward_tiu_cube_time"] = 0
    res_set[k]["grad_add_1684x_backward_ops_time"]      = res_set[k]["grad_add_1684x_backward_tiu_vec_time"] + res_set[k]["grad_add_1684x_backward_tiu_cube_time"]
    res_set[k]["grad_add_2260_backward_tiu_vec_time"]   = (res_set[k]["grad_chain_ops"] / params["2260 vector tiu"])    * 1 if res_set[k]["input_dtype"][0] == "float16" else rate
    res_set[k]["grad_add_2260_backward_tiu_cube_time"]  = 0
    res_set[k]["grad_add_2260_backward_ops_time"]       = res_set[k]["grad_add_2260_backward_tiu_vec_time"] + res_set[k]["grad_add_2260_backward_tiu_cube_time"]

def calc_all_time(res_set):
    total_times = {
        "1684x_forward" : 0,
        "2260_forward"  : 0,
        "1684x_backward": 0,
        "2260_backward" : 0,
        "1684x_fwd_dma" : 0,
        "2260_fwd_dma": 0,
        "1684x_fwd_tiu": 0,
        "1684x_fwd_tiu_vec": 0,
        "1684x_fwd_tiu_cube": 0,
        "2260_fwd_tiu": 0,
        "2260_fwd_tiu_vec": 0,
        "2260_fwd_tiu_cube": 0,
        "1684x_bwd_dma": 0,
        "2260_bwd_dma": 0,
        "1684x_bwd_tiu": 0,
        "1684x_bwd_tiu_vec": 0,
        "1684x_bwd_tiu_cube": 0,
        "2260_bwd_tiu": 0,
        "2260_bwd_tiu_vec": 0,
        "2260_bwd_tiu_cube": 0,
        "forward_ops": 0,
        "backward_ops": 0,
        "total_ops": 0,
        "total_dma": 0,
        "2260_mac_utils": 0,
    }
    ops_total_times = defaultdict(dict)
    for k in res_set.keys():
        if k not in need_calc_nodes and do_select:
            continue
        # forward time 
        name = res_set[k]["name"]
        res_set[k]["1684x_forward_dma_time"]   = res_set[k]["forward_dma"] / params["1684x dma"]
        res_set[k]["2260_forward_dma_time"]    = res_set[k]["forward_dma"] / params["2260 dma"]
        # res_set[k]["1684x_forward_ops_time"]   = res_set[k]["ops"] / (params["1684x f16 tiu"] if res_set[k]["input_dtype"][0] == "float16" else params["1684x f32 tiu"])
        # res_set[k]["2260_forward_ops_time"]    = res_set[k]["ops"] / (params["2260 f16 tiu"]  if res_set[k]["input_dtype"][0] == "float16" else params["2260 f32 tiu"])
        handle_tiu_time_calc(res_set, k)
        res_set[k]["1684x_forward_time"]       = max(res_set[k]["1684x_forward_dma_time"],res_set[k]["1684x_forward_ops_time"])
        res_set[k]["2260_forward_time"]        = max(res_set[k]["2260_forward_dma_time"],res_set[k]["2260_forward_ops_time"])
        total_times["1684x_forward"]           += res_set[k]["1684x_forward_time"]
        total_times["2260_forward"]            += res_set[k]["2260_forward_time"]
        total_times["1684x_fwd_dma"]           += res_set[k]["1684x_forward_dma_time"]
        total_times["2260_fwd_dma"]            += res_set[k]["2260_forward_dma_time"]
        total_times["1684x_fwd_tiu"]           += res_set[k]["1684x_forward_ops_time"]
        total_times["1684x_fwd_tiu_vec"]       += res_set[k]["1684x_forward_tiu_vec_time"]
        total_times["1684x_fwd_tiu_cube"]      += res_set[k]["1684x_forward_tiu_cube_time"]
        total_times["2260_fwd_tiu"]            += res_set[k]["2260_forward_ops_time"]
        total_times["2260_fwd_tiu_vec"]        += res_set[k]["2260_forward_tiu_vec_time"]
        total_times["2260_fwd_tiu_cube"]       += res_set[k]["2260_forward_tiu_cube_time"]
        total_times["forward_ops"]             += res_set[k]["ops"]
        total_times["total_ops"]               += res_set[k]["ops"]
        total_times["total_dma"]               += res_set[k]["forward_dma"]
        if "1684x_forward" not in ops_total_times[name]:
            ops_total_times[name]["1684x_forward"] = 0
            ops_total_times[name]["2260_forward"]  = 0
            ops_total_times[name]["1684x_dma"]     = 0
            ops_total_times[name]["2260_dma"]      = 0
            ops_total_times[name]["1684x_tiu"]     = 0
            ops_total_times[name]["1684x_tiu_vec"] = 0
            ops_total_times[name]["1684x_tiu_cube"]= 0
            ops_total_times[name]["2260_tiu"]      = 0
            ops_total_times[name]["2260_tiu_vec"]  = 0
            ops_total_times[name]["2260_tiu_cube"] = 0
            ops_total_times[name]["forward_ops"]   = 0
            ops_total_times[name]["forward_dma"]   = 0
            ops_total_times[name]["fwd_count"]     = 0
        ops_total_times[name]["1684x_forward"] += res_set[k]["1684x_forward_time"]
        ops_total_times[name]["2260_forward"]  += res_set[k]["2260_forward_time"]
        ops_total_times[name]["1684x_dma"]     += res_set[k]["1684x_forward_dma_time"]
        ops_total_times[name]["2260_dma"]      += res_set[k]["2260_forward_dma_time"]
        ops_total_times[name]["1684x_tiu"]     += res_set[k]["1684x_forward_ops_time"]
        ops_total_times[name]["1684x_tiu_vec"] += res_set[k]["1684x_forward_tiu_vec_time"]
        ops_total_times[name]["1684x_tiu_cube"]+= res_set[k]["1684x_forward_tiu_cube_time"]
        ops_total_times[name]["2260_tiu"]      += res_set[k]["2260_forward_ops_time"]
        ops_total_times[name]["2260_tiu_vec"]  += res_set[k]["2260_forward_tiu_vec_time"]
        ops_total_times[name]["2260_tiu_cube"] += res_set[k]["2260_forward_tiu_cube_time"]
        ops_total_times[name]["forward_ops"]   += res_set[k]["ops"]
        ops_total_times[name]["forward_dma"]   += res_set[k]["forward_dma"]
        ops_total_times[name]["fwd_count"]     += 1
        # backward time
        if k in activation_nodes:
            res_set[k]["1684x_backward_dma_time"] = res_set[k]["back_dma"] / params["1684x dma"]
            res_set[k]["2260_backward_dma_time"]  = res_set[k]["back_dma"] / params["2260 dma"]
            # res_set[k]["1684x_backward_ops_time"] = res_set[k]["back_ops"] / (params["1684x f16 tiu"] if res_set[k]["input_dtype"][0] == "float16" else params["1684x f32 tiu"])
            # res_set[k]["2260_backward_ops_time"]  = res_set[k]["back_ops"] / (params["2260 f16 tiu"]  if res_set[k]["input_dtype"][0] == "float16" else params["2260 f32 tiu"])
            handle_bwd_tiu_time_calc(res_set,k)
            res_set[k]["1684x_backward_time"]     = max(res_set[k]["1684x_backward_dma_time"],res_set[k]["1684x_backward_ops_time"])
            res_set[k]["2260_backward_time"]      = max(res_set[k]["2260_backward_dma_time"],res_set[k]["2260_backward_ops_time"])
            total_times["1684x_backward"]         += res_set[k]["1684x_backward_time"]
            total_times["2260_backward"]          += res_set[k]["2260_backward_time"]  
            total_times["1684x_bwd_dma"]          += res_set[k]["1684x_backward_dma_time"]
            total_times["2260_bwd_dma"]           += res_set[k]["2260_backward_dma_time"]
            total_times["1684x_bwd_tiu"]          += res_set[k]["1684x_backward_ops_time"]
            total_times["1684x_bwd_tiu_vec"]      += res_set[k]["1684x_backward_tiu_vec_time"]
            total_times["1684x_bwd_tiu_cube"]     += res_set[k]["1684x_backward_tiu_cube_time"]
            total_times["2260_bwd_tiu"]           += res_set[k]["2260_backward_ops_time"]
            total_times["2260_bwd_tiu_vec"]       += res_set[k]["2260_backward_tiu_vec_time"]
            total_times["2260_bwd_tiu_cube"]      += res_set[k]["2260_backward_tiu_cube_time"]
            total_times["backward_ops"]           += res_set[k]["back_ops"]
            total_times["total_ops"]              += res_set[k]["back_ops"] if "back_recompute" not in res_set[k] else res_set[k]["back_ops"] - res_set[k]["back_recompute"]
            total_times["total_dma"]              += res_set[k]["back_dma"]
            if "1684x_backward" not in ops_total_times[name]:
                ops_total_times[name]["1684x_backward"] = 0
                ops_total_times[name]["2260_backward"]  = 0
                ops_total_times[name]["1684x_bwd_dma"]  = 0
                ops_total_times[name]["2260_bwd_dma"]   = 0
                ops_total_times[name]["1684x_bwd_tiu"]  = 0
                ops_total_times[name]["1684x_bwd_tiu_vec"] = 0
                ops_total_times[name]["1684x_bwd_tiu_cube"]= 0
                ops_total_times[name]["2260_bwd_tiu"]   = 0
                ops_total_times[name]["2260_bwd_tiu_vec"] = 0
                ops_total_times[name]["2260_bwd_tiu_cube"]= 0
                ops_total_times[name]["backward_ops"]   = 0
                ops_total_times[name]["backward_dma"]   = 0
                ops_total_times[name]["bwd_count"]      = 0
                ops_total_times[name]["grad_add_1684x_backward_dma_time"] = 0
                ops_total_times[name]["grad_add_2260_backward_dma_time"]  = 0
                ops_total_times[name]["grad_add_1684x_backward_tiu_time"] = 0
                ops_total_times[name]["grad_add_2260_backward_tiu_time"]  = 0
                ops_total_times[name]["grad_add_count"] = 0
                ops_total_times[name]["grad_add_ops"]   = 0
                ops_total_times[name]["grad_add_dma"]   = 0
                ops_total_times[name]["grad_add_1684x_backward"] = 0
                ops_total_times[name]["grad_add_2260_backward"]  = 0
            ops_total_times[name]["1684x_backward"]     += max(res_set[k]["1684x_backward_ops_time"], res_set[k]["1684x_backward_dma_time"])
            ops_total_times[name]["2260_backward"]      += max(res_set[k]["2260_backward_ops_time"],  res_set[k]["2260_backward_dma_time"])
            ops_total_times[name]["1684x_bwd_dma"]      += res_set[k]["1684x_backward_dma_time"]
            ops_total_times[name]["2260_bwd_dma"]       += res_set[k]["2260_backward_dma_time"]
            ops_total_times[name]["1684x_bwd_tiu"]      += res_set[k]["1684x_backward_ops_time"]
            ops_total_times[name]["1684x_bwd_tiu_vec"]  += res_set[k]["1684x_backward_tiu_vec_time"]
            ops_total_times[name]["1684x_bwd_tiu_cube"] += res_set[k]["1684x_backward_tiu_cube_time"]
            ops_total_times[name]["2260_bwd_tiu"]       += res_set[k]["2260_backward_ops_time"]
            ops_total_times[name]["2260_bwd_tiu_vec"]   += res_set[k]["2260_backward_tiu_vec_time"]
            ops_total_times[name]["2260_bwd_tiu_cube"]  += res_set[k]["2260_backward_tiu_cube_time"]
            ops_total_times[name]["backward_ops"]       += res_set[k]["back_ops"]
            ops_total_times[name]["backward_dma"]       += res_set[k]["back_dma"]
            ops_total_times[name]["bwd_count"]          += 1
            if "grad_chain_count" in res_set[k]:
                if "grad_add_1684x_backward_dma_time" not in res_set[k]:
                    res_set[k]["grad_add_1684x_backward_dma_time"] = 0
                    res_set[k]["grad_add_2260_backward_dma_time"]  = 0
                    res_set[k]["grad_add_1684x_backward_ops_time"] = 0
                    res_set[k]["grad_add_2260_backward_ops_time"]  = 0
                res_set[k]["grad_add_1684x_backward_dma_time"]            += res_set[k]["grad_chain_dma"] / params["1684x dma"]
                res_set[k]["grad_add_2260_backward_dma_time"]             += res_set[k]["grad_chain_dma"] / params["2260 dma"]
                handle_grad_add_time_calc(res_set, k)
                # res_set[k]["grad_add_1684x_backward_ops_time"]            += res_set[k]["grad_chain_ops"] / (params["1684x f16 tiu"] if res_set[k]["input_dtype"][0] == "float16" else params["1684x f32 tiu"])
                # res_set[k]["grad_add_2260_backward_ops_time"]             += res_set[k]["grad_chain_ops"] / (params["2260 f16 tiu"]  if res_set[k]["input_dtype"][0] == "float16" else params["2260 f32 tiu"])
                total_times["total_ops"]                                  += res_set[k]["grad_chain_ops"]
                total_times["total_dma"]                                  += res_set[k]["grad_chain_dma"]
                total_times["1684x_backward"]                             += max(res_set[k]["grad_add_1684x_backward_dma_time"], res_set[k]["grad_add_1684x_backward_ops_time"])
                total_times["2260_backward"]                              += max(res_set[k]["grad_add_2260_backward_dma_time"], res_set[k]["grad_add_2260_backward_ops_time"])
                total_times["1684x_bwd_dma"]                              += res_set[k]["grad_add_1684x_backward_dma_time"]
                total_times["2260_bwd_dma"]                               += res_set[k]["grad_add_2260_backward_dma_time"]
                total_times["1684x_bwd_tiu"]                              += res_set[k]["grad_add_1684x_backward_ops_time"]
                total_times["1684x_bwd_tiu_vec"]                          += res_set[k]["grad_add_1684x_backward_tiu_vec_time"]
                total_times["1684x_bwd_tiu_cube"]                         += res_set[k]["grad_add_1684x_backward_tiu_cube_time"]
                total_times["2260_bwd_tiu"]                               += res_set[k]["grad_add_2260_backward_ops_time"]
                total_times["2260_bwd_tiu_vec"]                           += res_set[k]["grad_add_2260_backward_tiu_vec_time"]
                total_times["2260_bwd_tiu_cube"]                          += res_set[k]["grad_add_2260_backward_tiu_cube_time"]
                ops_total_times[name]["grad_add_1684x_backward"]          += max(res_set[k]["grad_add_1684x_backward_dma_time"], res_set[k]["grad_add_1684x_backward_ops_time"])
                ops_total_times[name]["grad_add_2260_backward"]           += max(res_set[k]["grad_add_2260_backward_dma_time"], res_set[k]["grad_add_2260_backward_ops_time"])
                ops_total_times[name]["grad_add_dma"]                     += res_set[k]["grad_chain_dma"]
                ops_total_times[name]["grad_add_ops"]                     += res_set[k]["grad_chain_ops"]
                ops_total_times[name]["grad_add_count"]                   += res_set[k]["grad_chain_count"]
                ops_total_times[name]["grad_add_1684x_backward_dma_time"] += res_set[k]["grad_add_1684x_backward_dma_time"]
                ops_total_times[name]["grad_add_2260_backward_dma_time"]  += res_set[k]["grad_add_2260_backward_dma_time"]
                ops_total_times[name]["grad_add_1684x_backward_tiu_time"] += res_set[k]["grad_add_1684x_backward_ops_time"]
                ops_total_times[name]["grad_add_2260_backward_tiu_time"]  += res_set[k]["grad_add_2260_backward_ops_time"]

    return res_set, total_times, ops_total_times

def write_to_sheet(ws, data, start_row, start_col):
    if isinstance(data, pd.DataFrame):
        for r_idx, row in enumerate(dataframe_to_rows(data, index=True, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)
        return r_idx  
    elif isinstance(data, list):
        for r_idx, line in enumerate(data, start=start_row):
            ws.cell(row=r_idx, column=start_col, value=line)
        return r_idx  
    else:
        raise ValueError("Unsupported data type")

params = {
    "1684x dma": 6e4,
    "2260 dma": 548e3,
    "1684x f16 tiu": 16e6,
    "1684x f32 tiu": 2e6,
    "1684x cuba tiu": 16e6,
    "1684x vector tiu": 4e6,
    "2260 f16 tiu": 128e6,
    "2260 f32 tiu": 16e6,
    "2260 cuba tiu": 128e6,
    "2260 vector tiu": 32e6,
}

batchs = [1]
whr = [512,768,960,1024,1280,1536,1792,2048]
# 如何更好生成表格
summary_table = defaultdict(dict)

for nb, nw in product(batchs, whr):
    nh = nw
    mylog.info("batch: %d, h: %d, w: %d"%(nb, nh, nw))
    cur_shapes = {
        "n": nb,
        "h": nh//8,
        "w": nw//8
    }
    shape_tuple = (nb, nh, nw)
    mylog.info("cur_shapes", cur_shapes)
    graph = deepcopy(source_graph)
    mylog.info(">>>>> walk for all activation mem usage")
    fix_shape_total_graph(cur_shapes)
    need_batch  = 16
    need_shapes = [512,768,960,1024]
    activation_mem_usage = walk_for_training_mem_activation_mem(graph)
    mylog.info(">>>>> make mem table for activation")
    # mem_table = pd.DataFrame.from_dict(activation_mem_usage, orient='index', columns=['op'])
    model_id = list(graph.keys())[0]
    model_graph = graph[model_id]["children"][0]
    model_graph_id = list(model_graph.keys())[0]
    graph_mem  = model_graph[model_graph_id]["graph_mem"]
    # mem_table["rate"] = mem_table["op"] / graph_mem
    # add total mem usage
    activation_mem_usage["total"] = sum(activation_mem_usage.values())
    mylog.info(">>>>> mem over")
    # forward and backward graph dma and ops calc
    res_set = forward_dma_ops_calc()
    res_set = prepare_bwd_node_outputs_calc(res_set)
    res_set = calc_bwd_bdc_dma(res_set)
    res_set = calc_bwd_grad_chain(res_set)
    # calc backward graph and memory usage(time)
    res_set, total_times, ops_times = calc_all_time(res_set)
    fwd_bwd_df = pd.DataFrame.from_dict(res_set, orient='index')
    # path -> padding into 6
    # in table add path_0, path_1, path_2, path_3, path_4, path_5
    # fwd_bwd_df["path_0"] = fwd_bwd_df["path"].apply(lambda x: x[0] if len(x) > 0 else "")
    fwd_bwd_df["path_1"] = fwd_bwd_df["path"].apply(lambda x: x[1] if len(x) > 1 else "")
    fwd_bwd_df["path_2"] = fwd_bwd_df["path"].apply(lambda x: x[2] if len(x) > 2 else "")
    fwd_bwd_df["path_3"] = fwd_bwd_df["path"].apply(lambda x: x[3] if len(x) > 3 else "")
    fwd_bwd_df["path_4"] = fwd_bwd_df["path"].apply(lambda x: x[4] if len(x) > 4 else "")
    # fwd_bwd_df["path_5"] = fwd_bwd_df["path"].apply(lambda x: x[5] if len(x) > 5 else "")
    remove_column_names = ["id", "path", "s2l_dma", "l2s_dma", "depth", "info","depth", "info","align_input_shape", "align_output_shape", "align_input_dtype", "align_output_dtype", "input_shape_layout", "output_shape_layout", "input_dtype", "output_dtype"]
    fwd_bwd_df = fwd_bwd_df.drop(remove_column_names, axis=1)
    # reorder columns
    new_names = ["path_1", "path_2", "path_3", "path_4", "name", "forward_dma", "ops", "1684x_forward_dma_time", "1684x_forward_ops_time", "1684x_forward_time", "2260_forward_dma_time", "2260_forward_ops_time", "2260_forward_time", "back_dma", "back_ops", "1684x_backward_dma_time", "1684x_backward_ops_time", "1684x_backward_time", "2260_backward_dma_time", "2260_backward_ops_time", "2260_backward_time", "leaf",  "cur_mem",  "back_mem"]
    fwd_bwd_df = fwd_bwd_df[new_names]
    # 去掉index
    fwd_bwd_df.reset_index(drop=True, inplace=True)
    fwd_bwd_df     = fwd_bwd_df.fillna(0)
    ops_times_df   = pd.DataFrame.from_dict(ops_times, orient='index').fillna(0)
    # add total time
    ops_times_df.loc["total"] = ops_times_df.sum()
    total_times["1684x_total_time"]           = total_times["1684x_forward"] + total_times["1684x_backward"]
    total_times["2260_total_time"]            = total_times["2260_forward"]  + total_times["2260_backward"]
    total_times["1684x_mac_utils"]            = total_times["total_ops"] * 100 / total_times["1684x_total_time"] / params["1684x f16 tiu"]
    total_times["2260_mac_utils"]             = total_times["total_ops"] * 100 / total_times["2260_total_time"] / params["2260 f16 tiu"]
    total_times["total_activation"]           = graph_mem
    summary_table[shape_tuple]["total"]       = total_times
    summary_table[shape_tuple]["ops_times"]   = ops_times_df
    summary_table[shape_tuple]["mem_table"]   = activation_mem_usage
    summary_table[shape_tuple]["fwd_bwd_df"]  = fwd_bwd_df

# summary tables
pre_table = {}
# total time summary table
all_shape_total_table = {}
all_shape_mem_table   = {}
# all_shape_ops_table   = {}
for k in summary_table.keys():
    kname = "_".join(str(i) for i in k)
    all_shape_total_table[kname] = summary_table[k]["total"]
    all_shape_mem_table[kname]   = summary_table[k]["mem_table"]


total_shape_tb     = pd.DataFrame(all_shape_total_table)
# 保存两位小数
total_shape_tb     = total_shape_tb.applymap(lambda x: round(x, 2))
# 排序 index 排序 
total_shape_tb     = total_shape_tb.sort_index()
total_shape_mem_tb = pd.DataFrame(all_shape_mem_table)
# each shape has a table 

wb = Workbook()
ws = wb.active
ws.title = "total_shape_tb"
annotation = ["For lora finetune, the mem usage and time usage of ADAM can be ignored. "]
annotation += ["time(us)                activation mem (B)"]
annotation += ["total time and total activation mem usage"]
last_row = write_to_sheet(ws, annotation, 5, 5)
last_row = write_to_sheet(ws, total_shape_tb, last_row+3, 5)

annotation = ["TABLE of Each Layer Activation mem usage"]
last_row = write_to_sheet(ws, annotation, last_row+3, 5)
last_row = write_to_sheet(ws, total_shape_mem_tb, last_row+3, 5)


for k in summary_table.keys():
    kname = "_".join(str(i) for i in k)
    ws = wb.create_sheet(title=kname)
    annotation = ["         TABLE of Each Layer Time usage     "]
    last_row = write_to_sheet(ws, annotation, 5, 5)
    op_times_df = summary_table[k]["ops_times"]
    last_row = write_to_sheet(ws, op_times_df, last_row+3, 5)
    annotation = ["         TABLE of Detail Layey Information      "]
    last_row = write_to_sheet(ws, annotation, last_row+3, 5)
    fwd_bwd_df = summary_table[k]["fwd_bwd_df"]
    last_row = write_to_sheet(ws, fwd_bwd_df, last_row+2, 5)

wb.save(output_file_name+".xlsx")
# total_shape_op_tb  = pd.DataFrame(all_shape_ops_table)


# import ipdb;ipdb.set_trace()


# # group by path_1, path_2, path_3, path_4

# import pdb;pdb.set_trace()